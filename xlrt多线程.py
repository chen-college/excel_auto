import openpyxl
import os
import time
import multiprocessing
import traceback
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

def check_chunk_for_empty_rows(args):
    """
    检查数据块中的空行（多进程工作函数）
    """
    file_path, sheet_name, start_row, end_row = args
    empty_rows = []
    
    try:
        # 每个进程单独打开工作簿
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook[sheet_name]
        
        # 获取数据范围（列范围）
        dims = sheet.calculate_dimension().split(':')
        if len(dims) < 2:
            return []  # 无数据
        
        # 使用标准方法解析坐标
        start_col, start_row_ref = coordinate_from_string(dims[0])
        end_col, end_row_ref = coordinate_from_string(dims[1])
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        # 只检查指定的行范围
        for row_idx in range(start_row, end_row + 1):
            is_row_empty = True
            for col_idx in range(start_col_idx, end_col_idx + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                # 判空逻辑
                if cell.value is not None:
                    if isinstance(cell.value, str) and cell.value.strip() == "":
                        continue
                    is_row_empty = False
                    break
            
            if is_row_empty:
                empty_rows.append(row_idx)
        
        workbook.close()
    except Exception as e:
        print(f"处理工作表 '{sheet_name}' 行 {start_row}-{end_row} 时出错: {str(e)}")
        traceback.print_exc()
    return empty_rows

def check_empty_rows_parallel(file_path, chunk_size=500):
    """
    使用多进程并行检测空行（修复Windows启动问题）
    """
    empty_rows_report = {}
    
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        # 只打开一次获取工作表名称
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()  # 立即关闭
        
        print(f"工作簿包含 {len(sheet_names)} 个工作表: {', '.join(sheet_names)}")
        
        # 获取Windows兼容的上下文
        ctx = multiprocessing.get_context('spawn')
        
        for sheet_name in sheet_names:
            print(f"\n处理工作表: '{sheet_name}'")
            
            try:
                # 获取工作表总行数
                workbook_temp = openpyxl.load_workbook(file_path, read_only=True)
                sheet = workbook_temp[sheet_name]
                dims = sheet.calculate_dimension().split(':')
                if len(dims) < 2:
                    print(f"  无数据，已跳过")
                    workbook_temp.close()
                    continue
                
                # 使用标准方法解析坐标
                start_col, start_row_ref = coordinate_from_string(dims[0])
                end_col, end_row_ref = coordinate_from_string(dims[1])
                start_row = int(start_row_ref)
                end_row = int(end_row_ref)
                total_rows = end_row - start_row + 1
                workbook_temp.close()
                
                # 准备分块处理（确保不重叠）
                chunks = []
                current = start_row
                while current <= end_row:
                    chunk_end = min(current + chunk_size - 1, end_row)
                    chunks.append((file_path, sheet_name, current, chunk_end))
                    current = chunk_end + 1  # 确保下一块不重叠
                
                # 使用进程池并行处理（使用spawn上下文）
                with ctx.Pool(processes=multiprocessing.cpu_count()) as pool:
                    results = pool.map(check_chunk_for_empty_rows, chunks)
                
                # 合并结果并去重
                empty_rows = []
                for result in results:
                    empty_rows.extend(result)
                empty_rows = sorted(set(empty_rows))  # 使用set去重
                
                # 打印发现的空行
                for row in empty_rows:
                    print(f"  发现空行: 第 {row} 行")
                
                if empty_rows:
                    empty_rows_report[sheet_name] = empty_rows
            except Exception as e:
                print(f"处理工作表 '{sheet_name}' 时发生严重错误: {str(e)}")
                traceback.print_exc()
    
    except Exception as e:
        print(f"处理Excel时发生全局错误: {str(e)}")
        traceback.print_exc()
        return {}
    
    return empty_rows_report

def main():
    """主函数：执行Excel检查操作并输出结果"""
    start_time = time.time()
    file_path = r'C:\Users\wjy17\Desktop\Excel_Scripts\#竞技场神兽配置表战力.xlsx'  # 修改为实际路径
    
    # 使用多进程检查空行（每500行一个块）
    empty_report = check_empty_rows_parallel(file_path, chunk_size=500)
    
    # 输出异常结果
    if empty_report:
        print("\n" + "="*50)
        print(">>> 异常报告: 空行统计 <<<")
        for sheet, rows in empty_report.items():
            print(f"工作表 '{sheet}': 共 {len(rows)} 个空行 - 行号: {rows}")
        print("="*50)
        
        # 导出结果到文件
        with open('空行报告.txt', 'w', encoding='utf-8') as f:
            f.write("空行检测报告\n")
            f.write("="*50 + "\n")
            for sheet, rows in empty_report.items():
                f.write(f"工作表 '{sheet}':\n")
                f.write(f"空行数量: {len(rows)}\n")
                f.write(f"行号列表: {', '.join(map(str, rows))}\n\n")
            print("检测结果已导出到: 空行报告.txt")
    else:
        print("\n>>> 检查结果: 所有工作表数据范围内无空行 <<<")
    
    # 计算并打印执行时间
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"\n优化后脚本执行时长: {execution_time:.4f} 秒")

if __name__ == "__main__":
    # Windows系统必需设置
    multiprocessing.freeze_support()
    multiprocessing.set_start_method('spawn', force=True)
    main()