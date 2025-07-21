import openpyxl
import os
import time
import traceback
from openpyxl.utils import get_column_letter

def detect_empty_cells(file_path):
    """
    检测Excel文件中所有工作表的空单元格
    返回字典格式 {工作表名: [空单元格坐标列表]}
    """
    empty_cells_report = {}
    
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        
        print(f"工作簿包含 {len(sheet_names)} 个工作表: {', '.join(sheet_names)}")
        
        for sheet_name in sheet_names:
            try:
                sheet = workbook[sheet_name]
                dims = sheet.calculate_dimension().split(':')
                if len(dims) < 2:
                    print(f"工作表 '{sheet_name}' 无数据，已跳过")
                    continue
                
                # 获取数据范围
                start_col, start_row = dims[0][0], int(dims[0][1:])
                end_col, end_row = dims[1][0], int(dims[1][1:])
                
                # 列字母转数字索引
                start_col_idx = openpyxl.utils.column_index_from_string(start_col)
                end_col_idx = openpyxl.utils.column_index_from_string(end_col)
                
                print(f"\n扫描工作表: '{sheet_name}' (数据范围: {dims[0]}-{dims[1]})")
                
                empty_cells = []
                
                # 遍历所有单元格
                for row_idx in range(start_row, end_row + 1):
                    for col_idx in range(start_col_idx, end_col_idx + 1):
                        cell = sheet.cell(row=row_idx, column=col_idx)
                        
                        # 判空逻辑
                        is_empty = False
                        if cell.value is None:
                            is_empty = True
                        elif isinstance(cell.value, str) and cell.value.strip() == "":
                            is_empty = True
                        elif cell.value == "":
                            is_empty = True
                        
                        # 记录空单元格
                        if is_empty:
                            col_letter = get_column_letter(col_idx)
                            cell_coord = f"{col_letter}{row_idx}"
                            empty_cells.append(cell_coord)
                            print(f"  发现空单元格: {cell_coord}")
                
                if empty_cells:
                    empty_cells_report[sheet_name] = empty_cells
            
            except Exception as e:
                print(f"\n处理工作表 '{sheet_name}' 时发生错误:")
                print(f"错误类型: {type(e).__name__}")
                print(f"错误信息: {str(e)}")
                traceback.print_exc()
        
        workbook.close()
    
    except Exception as e:
        print(f"处理Excel时发生错误: {str(e)}")
        traceback.print_exc()
        return {}
    
    return empty_cells_report

def main():
    start_time = time.time()
    file_path = r'C:\Users\wjy17\Desktop\Excel_Scripts\#竞技场神兽配置表战力.xlsx'
    
    # 检测空单元格
    empty_report = detect_empty_cells(file_path)
    
    # 输出结果
    if empty_report:
        print("\n" + "="*50)
        print(">>> 空单元格检测报告 <<<")
        for sheet, cells in empty_report.items():
            print(f"\n工作表 '{sheet}':")
            print(f"共发现 {len(cells)} 个空单元格")
            print("空单元格坐标:", ", ".join(cells))
        print("="*50)
        
        # 导出结果到文本文件
        with open('空单元格报告.txt', 'w', encoding='utf-8') as f:
            f.write("空单元格检测报告\n")
            f.write("="*50 + "\n")
            for sheet, cells in empty_report.items():
                f.write(f"\n工作表 '{sheet}':\n")
                f.write(f"空单元格数量: {len(cells)}\n")
                f.write("坐标列表:\n" + "\n".join(cells) + "\n")
            print("\n检测结果已导出到: 空单元格报告.txt")
    else:
        print("\n>>> 检查结果: 所有工作表未发现空单元格 <<<")
    
    end_time = time.time()
    execution_time = end_time - start_time
    print(f"\n脚本执行时长: {execution_time:.4f} 秒")

if __name__ == "__main__":
    main()